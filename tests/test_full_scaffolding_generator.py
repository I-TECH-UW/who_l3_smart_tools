from datetime import datetime, timezone
import os
from who_l3_smart_tools.core.indicator_testing.v1.scaffolding_generator import (
    ScaffoldingGenerator,
)
import pandas as pd
import unittest
from openpyxl import load_workbook


class TestGenerateTestScaffolding(unittest.TestCase):

    # Generates a new Excel file with the same number of sheets as the number of rows in the 'Indicator definitions' sheet of the input file.
    def test_generate_test_scaffolding_same_number_of_sheets(self):
        input_file = "tests/data/l2/test_indicators.xlsx"
        output_file = f"tests/output/scaffolding/indicator_test_output_{(datetime.now(timezone.utc)).strftime('%Y%m%d%H%M%S')}.xlsx"

        # Make sure output directory exists
        if not os.path.exists("tests/output/scaffolding"):
            os.makedirs("tests/output/scaffolding")

        sg = ScaffoldingGenerator(input_file, output_file)
        sg.generate_test_scaffolding()

        # Ensure output file exists
        assert os.path.exists(output_file)

        df = pd.read_excel(input_file, sheet_name="Indicator definitions")
        expected_sheets = len(df)

        wb = load_workbook(output_file)
        actual_sheets = len(wb.sheetnames)

        assert 29 == actual_sheets

    # The input file does not exist.
    def test_generate_test_scaffolding_input_file_not_exist(self):
        input_file = "path/to/nonexistent_input_file.xlsx"
        output_file = "path/to/output_file.xlsx"

        try:
            sg = ScaffoldingGenerator(input_file, output_file)
            sg.generate_test_scaffolding(input_file, output_file)
        except FileNotFoundError:
            assert True

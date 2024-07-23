import builtins
import os
from io import StringIO
from unittest import TestCase
from unittest.mock import mock_open, patch

from faker import Faker
from openpyxl import Workbook

from who_l3_smart_tools.core.logical_models.logical_model_generator import (
    LogicalModelGenerator,
)


class TestLogicalModelGenerator(TestCase):
    """
    A test case class for the LogicalModelGenerator class.
    """

    workbook_header = (
        "Activity ID",
        "Data Element ID",
        "Data Element Label",
        "Description and Definition",
        "Multiple Choice Type (if applicable)",
        "Data Type",
        "Input Options",
        "Quantity Sub-type",
        "Calculation",
        "Validation Condition",
        "Required",
        "Explain Conditionality",
    )

    def _create_dummy_workbook(self):
        workbook = Workbook()
        workbook.create_sheet(title="COVER")
        workbook.create_sheet(title="HIV.Sheet1")
        workbook.create_sheet(title="HIV.Sheet2")
        workbook.remove(workbook.active)  # Remove the default sheet
        return workbook

    def _add_cover_data(self, workbook):
        cover = workbook["COVER"]
        header = ["Sheet name", "Description"]
        cover.append(header)
        cover.append(["HIV.A Registration", "Sheet 1 description"])
        cover.append(["HIV.B Visit", "Sheet 2 description"])

    def _add_dummy_data(self, workbook, num_rows=20):
        fake = Faker()
        for sheet in workbook.sheetnames:
            if sheet.startswith("HIV"):
                ws = workbook[sheet]
                ws.append(self.workbook_header)
                while ws.max_row < num_rows:
                    row = [
                        fake.uuid4(),
                        fake.uuid4(),
                        fake.text(),
                        fake.text(),
                        fake.text(),
                        fake.random_element(
                            elements=[
                                "Boolean",
                                "String",
                                "Coding",
                                "DateTime",
                                "Date",
                                "ID",
                                "Quantity",
                            ]
                        ),
                        fake.text(),
                        fake.text(),
                        fake.text(),
                        fake.random_element(
                            elements=[
                                "Minimum and maximum number of characters based on country",
                                "Minimum and maximum number of characters based on local policy",
                                "Minimum and maximum number of characters, based on local policy",
                                "Must be appropriate email format with '@' sign",
                            ]
                        ),
                        fake.random_element(elements=["R", "NR"]),
                        fake.text(),
                    ]
                    ws.append(row)

    def setUp(self) -> None:
        self.wb_patcher = patch(
            "who_l3_smart_tools.core.logical_models.logical_model_generator.load_workbook"
        )
        self.mock_func = self.wb_patcher.start()
        self.mock_func.return_value = self._create_dummy_workbook()

        self.mock_written_files = {}
        self.original_open = builtins.open

        # This is a bit messy because we need to patch only writing but not reading
        def custom_open(
            file,
            mode="r",
            buffering=-1,
            encoding=None,
            errors=None,
            newline=None,
            closefd=True,
            opener=None,
        ):
            if mode.startswith("r"):
                # For reading, use the actual file system
                return self.original_open(
                    file, mode, buffering, encoding, errors, newline, closefd, opener
                )
            if mode.startswith("w") or mode.startswith("a"):
                # For writing or appending, use a StringIO object
                file_object = StringIO()
                mock_file_object = mock_open().return_value
                mock_file_object.write.side_effect = file_object.write
                if file not in self.mock_written_files:
                    self.mock_written_files[file] = file_object
                return mock_file_object
            raise ValueError(f"Unsupported file mode: {mode}")

        self.open_patcher = patch("builtins.open", custom_open)
        self.mock_open = self.open_patcher.start()

        self.logical_model_generator = LogicalModelGenerator("dummy.xlsx", "output")
        self._add_cover_data(self.logical_model_generator.workbook)
        self._add_dummy_data(self.logical_model_generator.workbook)

    def tearDown(self) -> None:
        self.wb_patcher.stop()
        self.open_patcher.stop()

    def test_process_cover_returns_sheet_info(self):
        cover_info = self.logical_model_generator.process_cover()
        self.assertDictEqual(
            cover_info,
            {
                "HIV.A REGISTRATION": "Sheet 1 description",
                "HIV.B VISIT": "Sheet 2 description",
            },
        )

    def test_clean_label_removes_unwanted_characters(self):
        label = 'Test*[]"'
        # pylint: disable=protected-access
        self.assertEqual(self.logical_model_generator._clean_label(label)[0], "Test'")

    def test_clean_label_returns_converted_label(self):
        label = "Test()'s-/, >=<=<>"
        # pylint: disable=protected-access
        self.assertEqual(
            self.logical_model_generator._clean_label(label)[1],
            "test___more thanless thanless thanmore than",
        )

    def test_map_cardinality_required_multichoice(self):
        cardinality = self.logical_model_generator.map_cardinality(
            "R", "Select all that apply"
        )
        self.assertEqual(cardinality, "1..*")

    def test_map_cardinality_multichoice(self):
        cardinality = self.logical_model_generator.map_cardinality(
            "", "Select all that apply"
        )
        self.assertEqual(cardinality, "0..*")

    def test_map_cardinality_required(self):
        cardinality = self.logical_model_generator.map_cardinality("R", "")
        self.assertEqual(cardinality, "1..1")

    def test_map_cardinality_not_required_not_multichoice(self):
        cardinality = self.logical_model_generator.map_cardinality("", "")
        self.assertEqual(cardinality, "0..1")

    def test_parse_validations_returns_all_validations(self):
        sheet_name = self.logical_model_generator.workbook.sheetnames[2]
        sheet = self.logical_model_generator.workbook[sheet_name]
        parsed_validations = self.logical_model_generator.parse_validations(sheet)
        expected = [
            "Minimum and maximum number of characters based on country",
            "Minimum and maximum number of characters based on local policy",
            "Minimum and maximum number of characters, based on local policy",
            "Must be appropriate email format with '@' sign",
        ]
        self.assertListEqual(list(parsed_validations.keys()), expected)

    def test_generate_fsh_from_excel_creates_relevant_files(self):
        self.logical_model_generator.generate_fsh_from_excel()
        files = [
            os.path.basename(filename) for filename in self.mock_written_files.keys()
        ]
        self.assertIn("HIVSheet1.fsh", files)
        self.assertIn("HIVSheet1.fsh", files)
        self.assertIn("HIVConcepts.fsh", files)

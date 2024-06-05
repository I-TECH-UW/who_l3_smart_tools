import unittest
import os
import pandas as pd
from openpyxl import load_workbook
from who_l3_smart_tools.core.indicator_testing.scaffolding_generator import (
    extract_elements,
    ScaffoldingGenerator,
)


@unittest.skip("Skip for now")
class TestExtractElements(unittest.TestCase):
    def setUp(self):
        self.test_cases = {
            'SUM of "Number of days prescribed" for all clients with "Medications prescribed"=\'PrEP for HIV prevention\'': {
                "expected_terms": {
                    "Number of days prescribed",
                    "Medications prescribed=PrEP for HIV prevention",
                },
                "expected_logic": "SUM of A for ALL WHERE B",
            },
            'COUNT of clients with "Medications prescribed"=\'PrEP for HIV prevention\' with "Date medications prescribed" in the reporting period': {
                "expected_terms": {
                    "Medications prescribed=PrEP for HIV prevention",
                    "Date medications prescribed in the reporting period",
                },
                "expected_logic": "COUNT WHERE A AND B",
            },
            'SUM of [DIFFERENCE in MIN("Date OAMT initiated", "Reporting period start date") and MAX("Date of loss to follow-up or OAMT stopped", "Reporting period end date")] for all clients with "Medications prescribed" IN \'Methadone\', \'Buprenorphine\'': {
                "expected_terms": {
                    "Date OAMT initiated",
                    "Reporting period start date",
                    "Date of loss to follow-up or OAMT stopped",
                    "Reporting period end date",
                    "Medications prescribed in Methadone, Buprenorphine",
                },
                "expected_logic": "SUM of [DIFFERENCE in MIN(A, B) and MAX(C, D)] for ALL WHERE E",
            },
            'COUNT of clients with ("Medications prescribed"=\'Methadone\' AND "Dose of medications prescribed" GREATER THAN OR EQUAL TO 60mg) OR ("Medications prescribed"=\'Buprenorphine\' AND "Dose of medications prescribed" GREATER THAN OR EQUAL TO 8mg) for a specified "Reporting date"': {
                "expected_terms": {
                    "Medications prescribed",
                    "Dose of medications prescribed",
                    "Reporting date",
                },
                "expected_logic": "COUNT of clients with (A='Methadone' AND B GREATER THAN OR EQUAL TO 60mg) OR (A='Buprenorphine' AND B GREATER THAN OR EQUAL TO 8mg) for a specified C",
            },
            'COUNT of clients with "HIV status"=\'HIV-positive\' AND "On ART"=True and "ART start date" GREATER THAN 6 months before reporting period end date AND "Date of viral load sample collection" within reporting period AND "Reason for HIV viral load test"=\'Routine viral load test\' AND "Viral load test result" LESS THAN 1000 copies/mL': {
                "expected_terms": {
                    "HIV status",
                    "On ART",
                    "ART start date",
                    "Date of viral load sample collection",
                    "Reason for HIV viral load test",
                    "Viral load test result",
                },
                "expected_logic": "COUNT of clients with A='HIV-positive' AND B=True and C GREATER THAN 6 months before reporting period end date AND D within reporting period AND E='Routine viral load test' AND F LESS THAN 1000 copies/mL",
            },
        }

    def test_extract_elements(self):
        for input_str, expected in self.test_cases.items():
            with self.subTest(input_str=input_str):
                terms, logic = extract_elements(input_str)
                self.assertEqual(
                    set(terms.keys()),
                    expected["expected_terms"],
                    f"Failed to extract terms for input: {input_str}",
                )
                self.assertEqual(
                    logic,
                    expected["expected_logic"],
                    f"Failed to construct logic for input: {input_str}",
                )


class ScaffoldingTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Define paths to your input and output Excel files
        cls.input_file = "tests/data/indicator_dak_input.xlsx"
        cls.output_file = "tests/output/indicator_test_output.xlsx"
        # Generate the test scaffolding output file
        # generate_test_scaffolding(cls.input_file, cls.output_file)

    @classmethod
    def tearDownClass(cls):
        # Remove the output file after tests are done
        # os.remove(cls.output_file)
        pass

    @unittest.skip("Skip this test for now")
    def test_sheet_columns(self):
        # Load the output workbook
        wb = load_workbook(self.output_file)
        df = pd.read_excel(self.input_file, sheet_name="Indicator definitions")

        for index, row in df.iterrows():
            ws_name = f"Indicator_{index + 1}"
            # Check if sheet exists
            self.assertIn(
                ws_name, wb.sheetnames, f"{ws_name} does not exist in the output file."
            )

            # Load the sheet
            ws = wb[ws_name]

            # Expected number of columns = # of terms + # of disaggregations + 4 (Patient ID, Logic String, Numerator, Denominator)
            expected_columns = (
                len(row["Disaggregation data elements"].split(","))
                + len(
                    row[
                        "List of all data elements included in numerator and denominator"
                    ].split(",")
                )
                + 4
            )
            actual_columns = ws.max_column

            self.assertEqual(
                expected_columns,
                actual_columns,
                f"{ws_name} has incorrect number of columns.",
            )


if __name__ == "__main__":
    unittest.main()

import os
from typing import Optional

from openpyxl import load_workbook

from who_l3_smart_tools.core.data_dictionary.utils import (
    remove_special_characters,
    to_camel_case,
)
from who_l3_smart_tools.utils.jinja2 import (
    DATA_TYPE_MAP,
    initalize_jinja_env,
    render_to_file,
)

jinja_env = initalize_jinja_env(__name__)


# pylint: disable=too-many-instance-attributes
class L2Row:
    """
    Represents a row in the L2 data dictionary.

    Args:
        raw_row (dict): The raw row data.
        coding_data_element (Optional[str]): The coding data element.

    Attributes:
        raw_row (dict): The raw row data.
        activity_id (str): The activity ID.
        data_element_id (Optional[str]): The data element ID.
        data_element_label (Optional[str]): The data element label.
        description (Optional[str]): The description and definition.
        choice_type (Optional[str]): The multiple choice type (if applicable).
        data_type (Optional[str]): The data type.
        input_options (Optional[str]): The input options.
        validation_condition (Optional[str]): The validation condition.
        required (Optional[str]): The required flag.
        coding_data_element (Optional[str]): The coding data element.

    Methods:
        validate_coding_data_element: Validates the coding data element.
        _activity_id_to_invariant: Converts the activity ID to an invariant.
        _get_questionare_title: Gets the questionnaire title.
        to_invariant: Converts the row to an invariant.
        to_concept_item: Converts the row to a concept item.
        to_model_item: Converts the row to a model item.
        to_questionnaire_item: Converts the row to a questionnaire item.
        to_valueset_item: Converts the row to a valueset item.
    """

    DEFAULT_INVARIANT_EXPRESSION = "<NOT-IMPLEMENTED>"
    DEFAULT_INVARIANT_SEVERITY = "error"

    def __init__(
        self, raw_row: dict, coding_data_element: Optional[str] = None
    ) -> None:
        self.raw_row = raw_row
        self.activity_id = raw_row["Activity ID"]
        self.data_element_id = raw_row["Data Element ID"]
        self.data_element_label = raw_row["Data Element Label"]
        self.description = raw_row["Description and Definition"]
        self.choice_type = raw_row["Multiple Choice Type (if applicable)"]
        self.data_type = raw_row["Data Type"]
        self.input_options = raw_row["Input Options"]
        self.validation_condition = raw_row["Validation Condition"]
        self.required = raw_row["Required"]
        self.coding_data_element = coding_data_element

    @property
    def map_cardinality(self):
        minimum = "0"
        maximum = "1"

        if self.required == "R":
            minimum = "1"

        if self.choice_type == "Select all that apply":
            maximum = "*"

        return f"{minimum}..{maximum}"

    def validate_coding_data_element(self) -> bool:
        if self.data_type == "Codes" and not self.coding_data_element:
            raise ValueError(
                f"Coding Data Element is required for data element {self.data_element_id}"
                f"of type Code"
            )

    @property
    def questionare_title(self) -> str:
        parts = self.activity_id.split(" ", 1)
        return parts[1] if len(parts) > 1 else parts[0]

    @property
    def question_instance(self) -> str:
        parts = self.activity_id.split(" ", 1)
        return (
            remove_special_characters(f"{parts[0]}{parts[1].capitalize()}")
            if len(parts) > 1
            else remove_special_characters(parts[0])
        )

    def to_invariant(self) -> Optional[dict[str, str]]:
        if self.validation_condition and self.validation_condition.lower() != "none":
            return {
                "id": self.activity_id[:5],
                "description": self.validation_condition,
                "expression": self.DEFAULT_INVARIANT_EXPRESSION,
                "severity": self.DEFAULT_INVARIANT_SEVERITY,
            }
        return None

    def to_concept_item(self) -> dict[str, str]:
        return {
            "id": self.data_element_id,
            "label": self.data_element_label,
            "description": self.description,
        }

    def to_model_item(self) -> dict[str, str]:
        return {
            "id": self.data_element_id,
            "slug": to_camel_case(self.data_element_label),
            "condition": self.map_cardinality,
            "type": DATA_TYPE_MAP[self.data_type],
            "label": self.data_element_label,
            "description": self.description,
        }

    def to_questionnaire_item(self) -> dict[str, str]:
        return {
            "id": self.data_element_id,
            "linkID": self.data_element_id,
            "type": self.data_type,
            "text": self.data_element_label,
            "required": "true" if self.required in ["R", "C"] else "false",
            "repeats": "false",
            "readOnly": "false",
        }

    def to_valueset_item(self) -> Optional[dict[str, str]]:
        if self.data_type == "Codes":
            return {"id": self.data_element_id, "label": self.data_element_label}
        return None

    def to_concept(self) -> dict[str, str]:
        return {
            "id": self.data_element_id,
            "label": self.data_element_label,
            "description": self.description,
        }


# pylint: disable=too-many-instance-attributes
class L2Dictionary:
    """
    Represents a data dictionary for Level 2 (L2) data.

    Args:
        file_path (str): The file path of the data dictionary.
        sheet_name_prefix (str, optional): The prefix of the sheet names to process.
        Defaults to "HIV".

    Attributes:
        workbook: The loaded workbook object.
        sheet_name_prefix (str): The prefix of the sheet names to process.
        active_coding_data_element: The currently active coding data element.
        concepts (list): A list of concepts extracted from the data dictionary.
        models (dict): A dictionary of models extracted from the data dictionary.
        questionnaires (dict): A dictionary of questionnaires extracted from the data dictionary.
        valuesets (dict): A dictionary of valuesets extracted from the data dictionary.

    Methods:
        set_active_coding(row): Sets the active coding data element based on the given row.
        add_to_model(sheet_name, row): Adds a row to the model based on the given sheet
            name and row.
        add_to_questionnaire(row): Adds a row to the questionnaire.
        add_to_valueset(row): Adds a row to the valueset.
        process(): Processes the data dictionary.

    """

    def __init__(
        self, file_path: str, output_path: str, sheet_name_prefix: str = "HIV"
    ) -> None:
        self.workbook = load_workbook(file_path)
        self.output_path = output_path
        self.sheet_name_prefix = sheet_name_prefix
        self.active_coding_data_element = None
        self.concepts = []
        self.models = {}
        self.questionnaires = {}
        self.valuesets = {}

    def set_active_coding(self, row: L2Row) -> None:
        if self.active_coding_data_element and row.data_type != "Codes":
            self.active_coding_data_element = None
        if row.data_type == "Coding":
            self.active_coding_data_element = row.data_element_id

    def add_to_model(self, sheet_name: str, row: L2Row) -> None:
        if row.data_type == "Codes":
            return
        _id = remove_special_characters(sheet_name)
        if _id in self.models:
            self.models[_id]["m_items"].append(row.to_model_item())
        else:
            self.models[_id] = {
                "m_items": [row.to_model_item()],
                "invariants": [],
                "title": sheet_name,
                "id": _id,
            }
        if row.to_invariant():
            row_invariant = row.to_invariant()
            max_id = max(
                [int(i["id"][6:]) for i in self.models[_id]["invariants"]] or [0]
            )
            row_invariant["id"] = f'{row_invariant["id"]}-{max_id + 1}'.replace(
                ".", "-"
            )
            invariant_texts = [i["description"] for i in self.models[_id]["invariants"]]
            if row_invariant["description"] not in invariant_texts:
                self.models[_id]["invariants"].append(row_invariant)

    def add_to_questionnaire(self, row: L2Row) -> None:
        if row.data_type == "Codes":
            return
        title = row.questionare_title
        if title in self.questionnaires:
            self.questionnaires[title]["q_items"].append(row.to_questionnaire_item())
        else:
            self.questionnaires[title] = {
                "q_items": [row.to_questionnaire_item()],
                "title": title,
                "instanceName": row.question_instance,
            }

    def add_to_valueset(self, row: L2Row) -> None:
        if row.data_type == "Codes":
            if self.active_coding_data_element in self.valuesets:
                self.valuesets[self.active_coding_data_element]["v_items"].append(
                    row.to_valueset_item()
                )
            else:
                self.valuesets[self.active_coding_data_element] = {
                    "v_items": [row.to_valueset_item()],
                    "name": remove_special_characters(self.active_coding_data_element),
                    "id": self.active_coding_data_element,
                }

    def process(self):
        for sheet_name in self.workbook.sheetnames:
            if not sheet_name.startswith(self.sheet_name_prefix):
                continue
            sheet = self.workbook[sheet_name]
            header: Optional[list[str]] = None
            for row in sheet.iter_rows(values_only=True):
                if not header:
                    header = row
                    continue
                raw_row = dict(zip(header, row))
                l2_row = L2Row(raw_row, self.active_coding_data_element)
                self.set_active_coding(l2_row)
                self.concepts.append(l2_row.to_concept())
                self.add_to_model(sheet_name, l2_row)
                self.add_to_questionnaire(l2_row)
                self.add_to_valueset(l2_row)

    def write_concepts(self):
        concepts_dir = "codesystems"
        output_path = os.path.join(self.output_path, concepts_dir, "HIVConcepts.fsh")
        os.makedirs(os.path.join(self.output_path, concepts_dir), exist_ok=True)
        template = jinja_env.get_template("concepts.fsh.j2")
        render_to_file(template, {"concepts": self.concepts}, output_path)

    def write_models(self):
        models_dir = "models"
        os.makedirs(os.path.join(self.output_path, models_dir), exist_ok=True)
        template = jinja_env.get_template("model.fsh.j2")
        for model in self.models.values():
            output_path = os.path.join(
                self.output_path, models_dir, f"{model['id']}.fsh"
            )
            render_to_file(template, {"model": model}, output_path)

    def write_questionnaires(self):
        questionnaires_dir = "questionnaires"
        os.makedirs(os.path.join(self.output_path, questionnaires_dir), exist_ok=True)
        template = jinja_env.get_template("questionnaire.fsh.j2")
        for questionnaire in self.questionnaires.values():
            output_path = os.path.join(
                self.output_path,
                questionnaires_dir,
                f"{questionnaire['instanceName']}.fsh",
            )
            render_to_file(template, {"q": questionnaire}, output_path)

    def write_valuesets(self):
        valuesets_dir = "valuesets"
        os.makedirs(os.path.join(self.output_path, valuesets_dir), exist_ok=True)
        template = jinja_env.get_template("valueset.fsh.j2")
        for valueset in self.valuesets.values():
            output_path = os.path.join(
                self.output_path, valuesets_dir, f"{valueset['name']}.fsh"
            )
            render_to_file(template, {"valueset": valueset}, output_path)

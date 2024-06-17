import pandas as pd
from itertools import product
from openpyxl import Workbook
import re
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Define fill colors
fills = {
    "Patient ID": PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    ),  # Yellow
    "Disaggregation": PatternFill(
        start_color="99FF99", end_color="99FF99", fill_type="solid"
    ),  # Light Green
    "Numerator Terms": PatternFill(
        start_color="FF9999", end_color="FF9999", fill_type="solid"
    ),  # Light Red
    "Denominator Terms": PatternFill(
        start_color="9999FF", end_color="9999FF", fill_type="solid"
    ),  # Light Blue
    "Other": PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    ),  # White
}


# Function to add unique items preserving order
def add_unique_preserving_order(original_list, items_to_add):
    for item in items_to_add:
        if item not in original_list:
            original_list.append(item)


def apply_background_color(ws, column_letter, fill):
    """
    Apply background color to an entire column.
    """
    for row in ws[column_letter]:
        row.fill = fill


def apply_fill_to_column_range(ws, start_col, end_col, fill):
    """
    Apply a background fill to a range of columns within a worksheet.
    """

    for col in range(start_col, end_col + 1):
        for row in ws[get_column_letter(col)]:
            row.fill = fill


def parse_exclusions(exclusions_str):
    try:
        pattern = r'with a[n]? "(.*?)" IN (.*?) at the end of the reporting period'
        match = re.search(pattern, exclusions_str)

        if match:
            variable = match.group(1)
            terms = match.group(2)
            terms_list = [term.strip() for term in terms.split(",")]
            result = [f'"{variable}" IN {term}' for term in terms_list]
            return exclusions_str, result
        else:
            return exclusions_str, exclusions_str.split(",")

    except Exception as e:
        print(f"Error parsing exclusions: {e}")
        return None, []


def parse_calculation(description):
    scope = "unknown"

    if description == "1":
        scope = "1"
        return [], scope

    client_scope_pattern = re.compile(r"of clients|for all clients", re.IGNORECASE)
    test_scope_pattern = re.compile(r"of tests|for all tests", re.IGNORECASE)

    if client_scope_pattern.search(description):
        scope = "clients"
    elif test_scope_pattern.search(description):
        scope = "tests"

    # Define the patterns
    count_pattern = re.compile(r"COUNT OF (.*?) WITH", re.IGNORECASE)
    sum_pattern = re.compile(r"SUM OF (.*?) FOR ALL CLIENTS WITH", re.IGNORECASE)
    term_pattern = re.compile(r"WITH (.*)", re.IGNORECASE)
    and_or_pattern = re.compile(r"\bAND\b|\bOR\b", re.IGNORECASE)
    op_split_terms = []

    # Find COUNT OF or SUM OF patterns
    count_match = count_pattern.search(description)
    sum_match = sum_pattern.search(description)

    if count_match:
        op_term_str = count_match.group(1)
    elif sum_match:
        op_term_str = sum_match.group(1)
        op_split_terms = re.split(
            r"\sAND\s|\sOR\s(?![^()]*\))", op_term_str, flags=re.IGNORECASE
        )
    else:
        return [], scope

    # Find the terms after the WITH
    term_match = term_pattern.search(description)
    if not term_match:
        return [], scope

    terms = term_match.group(1)

    # Split the terms by AND or OR, but not within parentheses
    split_terms = re.split(r"\sAND\s|\sOR\s(?![^()]*\))", terms, flags=re.IGNORECASE)

    return_terms = []
    if op_split_terms and len(op_split_terms) > 0:
        return_terms.extend(op_split_terms)
    return_terms.extend(split_terms)

    return_terms = [term.strip() for term in return_terms]

    return return_terms, scope


def extract_elements(calculation_str):
    # Extract the main operation (SUM or COUNT)
    operation_match = re.match(r"(SUM|COUNT)\s+", calculation_str)
    operation = operation_match.group(1) if operation_match else ""

    if not operation:
        return {}, 1

    # Initialize storage for unique terms
    unique_terms = set()
    sum_expression_terms = set()  # For terms inside a SUM expression

    # Handle complex SUM expressions separately
    if operation == "SUM":
        # Look for a term immediately following "SUM of" for simple SUM operations
        simple_sum_term_match = re.search(
            r'SUM of\s+"([^"]+)"\s+for all clients', calculation_str
        )
        if simple_sum_term_match:
            sum_expression_terms.add(simple_sum_term_match.group(1))

        # Handle complex SUM expressions involving DIFFERENCE, MIN, and MAX
        complex_sum_term_match = re.search(
            r"SUM of \[(.*?)\] for all clients", calculation_str, re.DOTALL
        )
        if complex_sum_term_match:
            sum_expression = complex_sum_term_match.group(1)
            # Extract terms from the complex SUM expression
            for term in re.findall(r'"([^"]+)"', sum_expression):
                sum_expression_terms.add(term)

    # Extract conditions outside of SUM expressions
    conditions_match = re.finditer(
        r'"([^"]+)"\s*(=|IN)\s*\'([^\']*)\'', calculation_str
    )
    for match in conditions_match:
        unique_terms.add(f"{match.group(1)} {match.group(2)} {match.group(3)}")

    # Handle "in the reporting period" phrases
    reporting_period_match = re.finditer(
        r'"([^"]+)"\s+in\s+the\s+reporting\s+period', calculation_str
    )
    for match in reporting_period_match:
        unique_terms.add(f"{match.group(1)} in the reporting period")

    # Combine and assign placeholders
    combined_terms = unique_terms.union(sum_expression_terms)
    term_to_placeholder = {term: chr(66 + i) for i, term in enumerate(combined_terms)}

    # Construct the logical expression
    return term_to_placeholder, calculation_str


class ScaffoldingGenerator:
    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file
        self.dak_data = pd.read_excel(input_file, sheet_name="Indicator definitions")
        self.writer = pd.ExcelWriter(output_file, engine="openpyxl")

    def generate_test_scaffolding(self):
        for index, row in self.dak_data.iterrows():
            if not (row["Included in DAK"] and row["Priority"] and row["Core"]):
                continue

            row_data = self.parse_dak_row(row)

            # Create a new worksheet for each indicator
            indicator_id = row_data["dak-id"]
            ws_name = f"{indicator_id}"
            workbook = self.writer.book

            if ws_name in workbook.sheetnames:
                ws = workbook[ws_name]
            else:
                ws = workbook.create_sheet(title=ws_name)

            numerator_scope, denominator_scope = (
                row_data["numerator-scope"],
                row_data["denominator-scope"],
            )

            if numerator_scope != denominator_scope:
                print(
                    f"Indicator {indicator_id} has different scopes for numerator and denominator calculations:\nnumerator:{numerator_scope} and denominator:{denominator_scope}"
                )

            headers, blank_fill_length = self.generate_headers(row, row_data)

            # Write headers to the worksheet
            ws.append(headers)

            # Generate true/false combinations for each term

            # Only create combinations for numerator and numerator exclusion, since
            # denominator should be a superset and might need to be cleaned up manually
            # due to the DAK imprecise definitions.
            combinations = list(
                product(
                    [0, 1],
                    repeat=(
                        len(row_data["numerator-terms"])
                        + len(row_data["numerator-exclusions"])
                    ),
                )
            )

            # If client scope, then each row represents a unique client. If test scope,
            # then each row represents a unique test. For simplicity, we only deal with the
            # scope indicated, and generate relevant patient data in later steps.
            for i, combo in enumerate(combinations, start=1):
                # Append combination rows with Patient ID and placeholders for numerator and denominator values
                ws.append(
                    [i] + [""] * blank_fill_length + list(combo)
                )  # Empty values for placeholder logic and actual num/denom values

            # Apply color fills based on calculated column ranges
            self.paint_worksheet(ws, row_data)

        self.writer.close()

    def parse_dak_row(self, row):
        # Extract the elements from the row
        row_data = {}

        row_data["short-name"] = row["Short name"]
        row_data["dak-id"] = row["DAK ID"]
        row_data["ref-id"] = row["Ref no."]

        # Extract disaggregation elements and data elements
        row_data["disaggregation-elements"] = re.split(
            r"[\n,]+", row["Disaggregation data elements"].strip()
        )

        # Extract Exclusion Elements if they exist and are strings
        row_data["numerator-exclusion-string"], row_data["numerator-exclusions"] = (
            parse_exclusions(row["Numerator exclusions"])
        )

        row_data["denominator-exclusion-string"], row_data["denominator-exclusions"] = (
            parse_exclusions(row["Denominator exclusions"])
        )

        # Parse the numerator and denominator calculations
        row_data["numerator-terms"], row_data["numerator-scope"] = parse_calculation(
            row["Numerator calculation"]
        )

        row_data["denominator-terms"], row_data["denominator-scope"] = (
            parse_calculation(row["Denominator calculation"])
        )

        return row_data

    def generate_headers(self, row, row_data):
        # Column 1
        if row_data["numerator-scope"] == "tests":
            headers = ["Test.id"]
        elif row_data["numerator-scope"] == "clients":
            headers = ["Patient.id"]
        else:
            print(
                f"Indicator {row_data['dak-id']} has an unknown scope: {row_data['numerator-scope']}"
            )
            headers = ["Patient.id"]

        # Construct headers, starting with 'Patient ID'
        headers = headers + ["Numerator:"] + [row["Numerator calculation"]]

        if row_data["numerator-exclusion-string"]:
            headers.append("EXCLUSION: " + row_data["numerator-exclusion-string"])

        blank_fill_length = len(headers) - 1

        headers = (
            headers + row_data["numerator-terms"] + row_data["numerator-exclusions"]
        )

        headers = headers + ["Denominator:"] + [row["Denominator calculation"]]

        if row_data["denominator-exclusion-string"]:
            headers.append("EXCLUSION: " + row_data["denominator-exclusion-string"])

        headers = (
            headers + row_data["denominator-terms"] + row_data["denominator-exclusions"]
        )

        headers = headers + ["Disaggregation:"] + row_data["disaggregation-elements"]

        return headers, blank_fill_length

    def paint_worksheet(self, ws, row_data):
        # Column 1 is always ID
        apply_fill_to_column_range(ws, 1, 1, fills["Patient ID"])

        # Numerator terms start after Patient ID
        num_start_col = 2
        num_end_col = (
            num_start_col
            + len(row_data["numerator-terms"])
            + len(row_data["numerator-exclusions"])
            + 1
        )
        if row_data["numerator-exclusion-string"]:
            num_end_col += 1

        apply_fill_to_column_range(
            ws, num_start_col, num_end_col, fills["Numerator Terms"]
        )

        # Denominator terms follow numerator terms
        denom_start_col = num_end_col + 1
        denom_end_col = (
            denom_start_col
            + len(row_data["denominator-terms"])
            + len(row_data["denominator-exclusions"])
            + 1
        )
        if row_data["denominator-exclusion-string"]:
            denom_end_col += 1
        apply_fill_to_column_range(
            ws, denom_start_col, denom_end_col, fills["Denominator Terms"]
        )

        # Disaggregation starts after
        dis_start_col = denom_end_col + 1
        dis_end_col = dis_start_col + len(row_data["disaggregation-elements"])
        apply_fill_to_column_range(
            ws, dis_start_col, dis_end_col, fills["Disaggregation"]
        )

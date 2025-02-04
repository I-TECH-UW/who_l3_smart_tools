import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import math


def generate_phenotype_xlsx(input_excel, output_excel, indicator=None):
    """
    For Domain Experts

    This function generates a template phenotype XLSX file based on the indicator definition XLSX file.

    It integrates with an interactive mode CLI command to load the excel file,
    list out available indicators, and prompt the user to select an indicator.
    It then uses the information from this indicator row to generate a template
    file with this information laid out in a way that assists domain experts with defining an exhaustive list of patient phenotypes for this indicator.
    """
    # Read the L2 indicator dataset
    df = pd.read_excel(input_excel, sheet_name="Indicator definitions")

    # Use DAK ID column for listing available indicators
    indicator_list = df["DAK ID"].unique()

    if indicator and indicator in indicator_list:
        selected_dak = indicator
    else:
        # Prompt user to select an indicator
        print("Available indicators:")
        for i, id_val in enumerate(indicator_list):
            print(f"{i}: {id_val}")
        selection = input("Select indicator (index or DAK id): ").strip()
        try:
            idx = int(selection)
            if idx < 0 or idx >= len(indicator_list):
                print("Invalid index selected.")
                return
            selected_dak = indicator_list[idx]
        except ValueError:
            if selection in indicator_list:
                selected_dak = selection
            else:
                print("Invalid DAK id selected.")
                return

    # Retrieve the selected indicator row
    indicator_row = df[df["DAK ID"] == selected_dak]
    if indicator_row.empty:
        print("Selected indicator not found in dataset.")
        return
    indicator_row = indicator_row.head(1)

    # Grab dissaggregation data elements
    disaggregation_elements = indicator_row["Disaggregation data elements"].values[0]
    # Split by newline or comma
    disaggregation_elements = [
        elem.strip() for elem in disaggregation_elements.replace("\n", ",").split(",")
    ]

    # Candidate headers for patient demographics and indicator features
    candidate_header = [
        "Patient Phenotype ID",
        "Phenotype Description",
    ]
    # Add disaggregation elements to candidate header
    candidate_header += disaggregation_elements
    # Add indicator feature columns
    candidate_header += [
        "<Add indicator feature columns>",
        "Counted as Numerator (0,1)",
        "Counted as Denominator (0,1)",
    ]

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        # Subset the indicator info to specified columns
        cols_subset = [
            "DAK ID",
            "Short name",
            "Indicator definition",
            "Numerator calculation",
            "Denominator calculation",
            "Denominator definition",
            "Disaggregation data elements",
            "Disaggregation description",
            "Numerator definition",
            "Numerator exclusions",
            "Denominator exclusions",
            "Ref no.",
            "List of all data elements included in numerator and denominator",
            "Comments and references",
            "Reference",
            "Category",
            "What it measures",
            "Rationale",
            "Method of measurement",
        ]
        indicator_subset = indicator_row[cols_subset]

        # Write three rows:
        # Row 1: indicator info header (column names)
        pd.DataFrame([cols_subset]).to_excel(
            writer, index=False, header=False, startrow=0
        )
        # Row 2: indicator info values
        indicator_subset.to_excel(writer, index=False, header=False, startrow=1)

        # Skip a row between indicator info and candidate header
        sheet_name = list(writer.sheets.keys())[0]
        writer.sheets[sheet_name].row_dimensions[3].height = 15

        # Row 4: candidate header row
        pd.DataFrame([candidate_header]).to_excel(
            writer, index=False, header=False, startrow=3
        )

        # Access the workbook and active worksheet
        workbook = writer.book
        worksheet = writer.sheets[workbook.sheetnames[0]]

        # Style rows:
        # Row 1 (Excel row 1): indicator header - dark yellow and bold
        dark_yellow_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        # Row 2 (Excel row 2): indicator values - light yellow
        light_yellow_fill = PatternFill(
            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
        )

        for cell in worksheet[1]:
            cell.fill = dark_yellow_fill
            cell.font = Font(bold=True)
        for cell in worksheet[2]:
            cell.fill = light_yellow_fill
        # Row 4 (Excel row 4): candidate header - bold text
        for cell in worksheet[4]:
            cell.font = Font(bold=True)

        # Define minimum and maximum column widths (adjust these values as needed)
        MIN_WIDTH = 10
        MAX_WIDTH = 50

        # Auto-resize columns based on max length in each column,
        # and enable word wrap if content exceeds the maximum width.
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            # Use the longest line among newline-split parts of each cell
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    lines = cell_value.split("\n")
                    longest_line = max(lines, key=len) if lines else ""
                    max_length = max(max_length, len(longest_line))
                except Exception:
                    pass
            # Compute adjusted width with a little buffer
            adjusted_width = max_length + 2
            # Set to minimum if too small...
            if adjusted_width < MIN_WIDTH:
                adjusted_width = MIN_WIDTH
            # ...or cap it at maximum and enable wrap_text if too wide.
            if adjusted_width > MAX_WIDTH:
                adjusted_width = MAX_WIDTH
            for cell in column_cells:
                current_alignment = cell.alignment or Alignment()
                cell.alignment = Alignment(
                    wrap_text=True,
                    horizontal=current_alignment.horizontal,
                    vertical=current_alignment.vertical,
                )
            worksheet.column_dimensions[column].width = adjusted_width

        # Adjust row heights to better display wrapped text.
        for row in worksheet.iter_rows():
            max_line_count = 1
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    # Split by existing newlines
                    lines = cell_value.split("\n")
                    total_lines = 0
                    # Get column width to estimate wrapping.
                    col_letter = cell.column_letter
                    col_width = (
                        worksheet.column_dimensions[col_letter].width or MAX_WIDTH
                    )
                    # Estimate how many lines are needed per line of text.
                    for line in lines:
                        # Use col_width as an approximation of how many characters are visible.
                        wrapped_lines = (
                            math.ceil(len(line) / col_width) if col_width else 1
                        )
                        total_lines += wrapped_lines
                        max_line_count = max(max_line_count, total_lines)
            # Adjust row height (default base height assumed 15, modify multiplier as needed)
            worksheet.row_dimensions[row[0].row].height = 15 * max_line_count

    print(f"Phenotype file generated: {output_excel}")

    return output_excel

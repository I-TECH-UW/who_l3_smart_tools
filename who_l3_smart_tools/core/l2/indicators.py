import os
from openpyxl import load_workbook
import stringcase

from who_l3_smart_tools.core.l2.data_dictionary import L2Dictionary
from who_l3_smart_tools.utils.jinja2 import initalize_jinja_env, render_to_file

jinja_env = initalize_jinja_env(__name__)


# pylint: disable=too-many-instance-attributes
# pylint: disable=no-member
class IndicatorRow:
    """
    Class representing an indicator row.

    Attributes:
        raw_row (dict): The raw row data.
        dak_id (str): The DAK ID.
        library_name (str): The library name.
        ref_no (str): The reference number.
        denominator (str): The denominator calculation.
        all_data_elements (list): The list of all data elements.
        included_in_dak (bool): Indicates if the indicator is included in DAK.

    Methods:
        set_other_columns_as_attributes(self) -> None:
            Sets the other columns as attributes.

        determine_scoring_suggestion(self) -> str:
            Determines the scoring suggestion based on the denominator value.

        to_cql(self) -> str:
            Converts the indicator to CQL format.
    """

    scoring_measure_instance = {
        "proportion": "http://hl7.org/fhir/us/cqfmeasures/StructureDefinition/proportion-measure-cqfm",
        "continuous-variable": "http://hl7.org/fhir/us/cqfmeasures/StructureDefinition/cv-measure-cqfm",
    }

    measure_required_elements = {
        "proportion": ["initialPopulation", "Numerator", "Denominator"],
        "continuous-variable": [
            "initialPopulation",
            "measurePopulation",
            "measureObservation",
        ],
    }

    def __init__(
        self, raw_row: dict, out_dir: str, linkage_concepts: dict[str, list]
    ) -> None:
        self.raw_row = raw_row
        self.out_dir = out_dir
        self.dak_id = raw_row.pop("DAK ID")
        self.library_name = f'{self.dak_id.replace(".", "")}Logic'
        self.ref_no = raw_row.pop("Ref no.")
        self.all_data_elements = (
            raw_row.pop(
                "List of all data elements included in numerator and denominator", ""
            )
            or ""
        ).split("\n")
        self.included_in_dak = raw_row.pop("Included in DAK")
        self.data_concepts = linkage_concepts.get(self.ref_no, [])
        self.set_other_columns_as_attributes()

    @property
    def disaggregations(self) -> str:
        return self.disaggregation_description.split("|")

    @property
    def scoring_method(self) -> str:
        return self.determine_scoring_suggestion()

    @property
    def proportion(self) -> bool:
        return self.scoring_method == "proportion"

    @property
    def continuous_variable(self) -> bool:
        return self.scoring_method == "continuous-variable"

    @property
    def scoring_instance(self) -> str:
        return self.scoring_measure_instance[self.scoring_method]

    def set_other_columns_as_attributes(self) -> None:
        for key, value in self.raw_row.items():
            key = stringcase.snakecase(key.lower())
            value = value.replace("\n", "|") if isinstance(value, str) else value
            setattr(self, key, value)

    def determine_scoring_suggestion(self) -> str:
        if (
            not self.denominator_calculation.strip()
            or self.denominator_calculation.strip() == "1"
        ):
            return "continuous-variable"
        return "proportion"

    def to_cql(self) -> str:
        output_path = os.path.join(self.out_dir, f"{self.library_name}.cql")
        template = jinja_env.get_template("indicators/cql_library.cql.j2")
        render_to_file(template, {"indicator": self}, output_path)


# pylint: disable=too-few-public-methods
class IndicatorLibrary:
    """
    A class representing an indicator library.

    Parameters:
    - indicator_file (str): The file path of the indicator file.
    - output_dir (str): The directory where the output will be saved.
    - data_dictionary_file (str): The file path of the data dictionary file.
    - sheet_name (str, optional): The name of the sheet in the indicator file.
    Defaults to "Indicator definitions".

    Methods:
    - generate_cql_scaffolds(): Generates CQL scaffolds based on the indicator definitions.
    """

    def __init__(
        self,
        indicator_file: str,
        output_dir: str,
        data_dictionary_file: str,
        sheet_name: str = "Indicator definitions",
    ) -> None:
        self.sheet = load_workbook(indicator_file)[sheet_name]
        self.data_dictionary = L2Dictionary(data_dictionary_file, output_dir)
        self.output_dir = output_dir

    def generate_cql_scaffolds(self) -> None:
        self.data_dictionary.process()
        header = None
        x = 0
        for row in self.sheet.iter_rows(values_only=True):
            if not header:
                header = row
                continue
            row = dict(zip(header, row))
            indicator_row = IndicatorRow(
                row, self.output_dir, self.data_dictionary.linkage_concepts
            )
            indicator_row.to_cql()
            print(
                indicator_row.numerator_exclusions,
                type(indicator_row.numerator_exclusions),
            )

            x += 1
            if x == 2:
                break

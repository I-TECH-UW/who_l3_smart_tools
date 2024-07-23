import os
import re
import sys
from collections import defaultdict

import inflect
import pandas as pd
import stringcase
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

from who_l3_smart_tools.utils import camel_case
from who_l3_smart_tools.utils.counter import Counter
from who_l3_smart_tools.utils.jinja2 import render_to_file

# TODO: differentiate between Coding, code, and CodableConcept
# Boolean
# String
# Date
# Time
# DateTime
# ID
# Quantity
# Signature
# Attachment
# Coding
# Codes
data_type_map = {
    "Boolean": "boolean",
    "String": "string",
    "Date": "date",
    "DateTime": "dateTime",
    "Coding": "Coding",
    "ID": "Identifier",
    "Quantity": "integer",
}

template_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")

jinja2_env = Environment(
    loader=FileSystemLoader(template_dir),
    trim_blocks=True,
    lstrip_blocks=True,
)


inflect_engine = inflect.engine()


# pylint: disable=too-many-locals,too-few-public-methods
class LogicalModelAndTerminologyGenerator:
    """
    Class for generating FSH logical models and terminologies based on an input Excel file.

    Attributes:
        input_file (str): The path to the input Excel file.
        output_dir (str): The directory where the generated FSH files will be saved.
        models_dir (str): The directory within the output directory where the
            logical models will be saved.
        codesystem_dir (str): The directory within the output directory
            where the code systems will be saved.
        valuesets_dir (str): The directory within the output directory
            where the value sets will be saved.
    """

    def __init__(self, input_file, output_dir):
        self.input_file = input_file
        self.output_dir = output_dir
        self.models_dir = os.path.join(output_dir, "models")
        self.codesystem_dir = os.path.join(output_dir, "codesystems")
        self.valuesets_dir = os.path.join(output_dir, "valuesets")
        self.invariants_dict = defaultdict(Counter)
        self.active_valueset = None

    # pylint: disable=too-many-branches,too-many-statements
    def generate_fsh_from_excel(self):
        """
        Generates FSH logical models and terminologies from the input Excel file.
        """
        for _dir in [self.models_dir, self.codesystem_dir, self.valuesets_dir]:
            if not os.path.exists(_dir):
                os.makedirs(_dir)

        # Load the Excel file
        workbook = load_workbook(self.input_file)

        # Process the Cover sheet
        cover_info = self.process_cover(workbook["COVER"])

        # Code system name
        code_system = "HIVConcepts"
        # store the actual codes as we process them
        codes = []

        validation_lookup = {}

        # Iterate over each sheet in the Excel file and generate a FSH logical model for each one

        # pylint: disable=too-many-nested-blocks
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("HIV."):
                sheet = workbook[sheet_name]

                # hard-coded, but the page labelled E-F has no F codes
                if sheet_name == "HIV.E-F PMTCT":
                    cleaned_sheet_name = "HIV.E PMTCT"
                else:
                    cleaned_sheet_name = sheet_name

                clean_name = stringcase.alphanumcase(cleaned_sheet_name)
                short_name = (cleaned_sheet_name.split(" ")[0]).split(".")

                # Initialize any ValueSets
                valuesets = []

                # Track element names
                existing_elements = defaultdict(Counter)

                # For handling "Other (specify)"
                previous_element_label = None

                # Process Invariants First
                # Get all unique validation conditions, and store their assigned rows
                validations = self.parse_validations(sheet)

                # Template for invariants based on validation conditions
                invariants = self.format_invariants(
                    validations, short_name, validation_lookup
                )

                self.render_elements(
                    sheet,
                    sheet_name,
                    cover_info,
                    code_system,
                    invariants,
                    codes,
                    valuesets,
                    existing_elements,
                    validation_lookup,
                    previous_element_label,
                    clean_name,
                )

        if codes:
            self.render_codes(codes, code_system)

    ### Helpers
    def process_cover(self, cover):
        """
        Process the cover sheet of the logical model.

        Args:
            cover (Worksheet): The cover sheet of the logical model.

        Returns:
            dict: A dictionary containing the processed cover data.
        """
        cover_data = {}

        seen_header = False
        for row in cover.iter_rows(values_only=True):
            if not seen_header:
                if (
                    row[0]
                    and isinstance(row[0], str)
                    and re.match(r"sheet\s*name", row[0], re.IGNORECASE)
                ):
                    seen_header = True
                continue

            if isinstance(row[0], str) and row[0]:
                key = row[0].upper()
                first_dot_idx = key.find(".")
                if len(key) > first_dot_idx >= 0:
                    if key[first_dot_idx + 1].isspace():
                        key = (
                            key[0:first_dot_idx]
                            + "."
                            + key[first_dot_idx + 1 :].lstrip()
                        )

                cover_data[key] = row[1]
            else:
                break

        return cover_data

    def map_cardinality(self, required_indicator, multiple_choice):
        """
        Maps the required indicator and multiple choice values to a cardinality string.

        Args:
            required_indicator (str): The required indicator value.
            multiple_choice (str): The multiple choice value.

        Returns:
            str: The cardinality string.

        """
        minimum = "0"
        maximum = "1"

        if required_indicator == "R":
            minimum = "1"

        if multiple_choice == "Select all that apply":
            maximum = "*"

        return f"{minimum}..{maximum}"

    def parse_validations(self, sheet):
        """
        Parses the validations from a given sheet and returns a dictionary of
            validation conditions and data element IDs.

        Args:
            sheet (pandas.DataFrame): The sheet containing the validations.

        Returns:
            dict: A dictionary where the keys are the validation conditions and the
                values are lists of data element IDs.
        """
        sheet_data = list(sheet.values)
        df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
        valids = df.groupby("Validation Condition")["Data Element ID"].groups
        return valids

    def _clean_label(self, label):
        """
        Cleans the label by removing special characters and converting it to lowercase.

        Args:
            label (str): The label to be cleaned.

        Returns:
            str: The cleaned label.
        """
        label = (
            label.strip()
            .replace("*", "")
            .replace("[", "")
            .replace("]", "")
            .replace('"', "'")
        )
        return label, (
            label.replace("(", "")
            .replace(")", "")
            .replace("'s", "")
            .replace("-", "_")
            .replace("/", "_")
            .replace(",", "")
            .replace(" ", "_")
            .replace(">=", "more than")
            .replace("<=", "less than")
            .replace(">", "more than")
            .replace("<", "less than")
            .lower()
        )

    # pylint: disable=too-many-arguments
    def _handle_value_sets(
        self,
        data_type,
        multiple_choice_type,
        data_element_id,
        label,
        description,
        code_system,
        valuesets,
    ):
        """
        Handles the creation of ValueSets and their codes.
        """
        # Originally, this looked at the Multiple Choice Type,
        # but that doesn't seem to be
        # guaranteed to be meaningful
        if data_type == "Coding":
            self.active_valueset = {
                "value_set": data_element_id,
                "name": data_element_id.replace(".", ""),
                "title": f"{label} ValueSet",
                "description": f"Value set of "
                f"{description[0].lower() + description[1:] \
                    if description[0].isupper() \
                    and not description.startswith("HIV") else description}",
                "codes": [],
            }
            valuesets.append(self.active_valueset)
        # Then we identify the codes for the ValueSet
        elif data_type == "Codes" and multiple_choice_type == "Input Option":
            if self.active_valueset is None:
                print(
                    f"Attempted to create a member of a ValueSet without a "
                    f"ValueSet context for code {data_element_id}",
                    sys.stderr,
                )
            else:
                self.active_valueset["codes"].append(
                    {
                        "code": f"{code_system}#{data_element_id}",
                        "label": f"{label}",
                    }
                )
        else:
            self.active_valueset = None

    # pylint: disable=too-many-arguments
    def render_elements(
        self,
        sheet,
        sheet_name,
        cover_info,
        code_system,
        invariants,
        codes,
        valuesets,
        existing_elements,
        validation_lookup,
        previous_element_label,
        clean_name,
    ):
        elements = []
        header = None
        for row in sheet.iter_rows(values_only=True):
            if header is None:
                header = row
                continue
            row = dict(zip(header, row))
            data_element_id = row["Data Element ID"]
            if not isinstance(data_element_id, str) or not data_element_id:
                continue

            # Process general fields
            multiple_choice_type = row["Multiple Choice Type (if applicable)"]
            data_type = row["Data Type"]
            label = row["Data Element Label"]

            if isinstance(label, str) and label:
                # Other (specify) elements come after a list as a data element to
                # contain a non-coded selection
                if label.lower() == "other (specify)":
                    if previous_element_label:
                        label_clean = f"Other_{previous_element_label.lower()}"
                    else:
                        label_clean = "Other (specify)"
                else:
                    label, label_clean = self._clean_label(label)
            else:
                label = ""
                label_clean = ""

            code_sys_ref = f"{code_system}#{data_element_id}"
            description = row["Description and Definition"]

            if isinstance(description, str):
                description = description.replace("*", "").replace('"', "'")
            else:
                description = ""

            required = row["Required"]

            codes.append(
                {
                    "code": data_element_id,
                    "label": label,
                    "description": description,
                }
            )

            self._handle_value_sets(
                data_type,
                multiple_choice_type,
                data_element_id,
                label,
                description,
                code_system,
                valuesets,
            )

            # If row is a value in a valueset, skip since the info is in Terminology
            if data_type == "Codes":
                continue

            # For any non-code we set the previous element name. This is used to determine the
            # name for "Other (specify)" data elements
            previous_element_label = label_clean

            # The camel-case version of the label becomes the element name in the logical model
            label_camel = camel_case(label_clean)

            # valid element identifiers in FHIR must start with a alphabetical character
            # therefore if an element starts with a number, we swap the spelt-out version of the
            # number, using the inflect library
            if len(label_camel) > 0 and not label_camel[0].isalpha():
                try:
                    prefix, rest = re.split(r"(?=[a-zA-Z])", label_camel, 1)
                except Exception:  # pylint: disable=broad-exception-caught
                    prefix, rest = label_camel, ""

                if prefix.isnumeric():
                    prefix = camel_case(
                        inflect_engine.number_to_words(int(prefix)).replace("-", "_")
                    )
                else:
                    print(
                        "Did not know how to handle element prefix:",
                        sheet_name,
                        data_element_id,
                        prefix,
                        file=sys.stderr,
                    )

                label_camel = f"{prefix}{rest}"

            # data elements can only be 64 characters
            # note that the idea here is that we trim whole words until reaching the desired size
            if len(label_camel) > 64:
                new_label_camel = ""
                for label_part in re.split("(?=[A-Z1-9])", label_camel):
                    if len(new_label_camel) + len(label_part) > 64:
                        break

                    new_label_camel += label_part
                label_camel = new_label_camel

            # data elements names must be unique per logical model
            count = existing_elements[label_camel].next

            # we have a duplicate data element
            if count > 1:
                # the first element needs no suffix
                # so the suffix is one less than the count
                suffix = str(count - 1)

                # if the data element id will still be less than 64 characters, we're ok
                if len(label_camel) + len(suffix) <= 64:
                    label_camel += suffix
                # otherwise, shorten the name to include the suffix
                else:
                    label_camel = label_camel[: 64 - len(suffix)] + suffix

            elements.append(
                {
                    "element_name": label_camel,
                    "cardinality": self.map_cardinality(required, multiple_choice_type),
                    "data_type": data_type_map[data_type],
                    "label": label,
                    "description": description,
                    "code": code_sys_ref,
                    "validation_id": validation_lookup.get(data_element_id),
                    "valueset": (
                        {"label": label_camel, "valueset": f"{data_element_id}"}
                        if data_type == "Coding"
                        else None
                    ),
                }
            )

        output_file = os.path.join(
            self.models_dir, f"{stringcase.alphanumcase(sheet_name)}.fsh"
        )

        template = jinja2_env.get_template("fsh_lm.j2")
        render_to_file(
            template,
            {
                "elements": elements,
                "invariants": invariants,
                "clean_name": clean_name,
                "sheet_name": sheet_name,
                "description": cover_info.get(sheet_name, ""),
            },
            output_file,
        )

        self.render_valuesets(valuesets)

    def format_invariants(self, validations, short_name, validation_lookup):
        """
        Formats the invariants based on the given validations, short name,
            and validation lookup.

        Args:
            validations (dict): A dictionary containing the validations.
            short_name (tuple): A tuple containing the short name.
            validation_lookup (dict): A dictionary containing the validation lookup.

        Returns:
            list: A list of formatted invariants.

        """
        invariants = []
        for validation, data_ids in validations.items():
            _id = self.invariants_dict[short_name[1]].next
            invariant_id = f"{short_name[0]}-{short_name[1]}-{_id}"
            if isinstance(validation, str):
                description = validation.replace('"', "'")
            else:
                description = ""

            invariants.append(
                {
                    "id": invariant_id,
                    "description": description,
                }
            )

            for data_element_id in data_ids:
                validation_lookup[data_element_id] = invariant_id
        return invariants

    def render_codes(self, codes, code_system):
        """
        Renders the codes to a code system file.

        Args:
            codes (list): A list of codes.
            code_system (str): The code system.

        Returns:
            None
        """
        title = ("WHO SMART HIV Concepts CodeSystem",)
        description = (
            (
                "This code system defines the concepts used in the"
                "World Health Organization SMART HIV DAK"
            ),
        )

        code_system_output_file = os.path.join(self.codesystem_dir, "HIVConcepts.fsh")
        template = jinja2_env.get_template("fsh_cs_codes.j2")
        render_to_file(
            template=template,
            context={
                "codes": codes,
                "title": title,
                "description": description,
                "code_system": code_system,
            },
            output_file=code_system_output_file,
        )

    def render_valuesets(self, valuesets):
        for valueset in valuesets:
            output_file = os.path.join(
                self.valuesets_dir, f"{valueset['value_set']}.fsh"
            )
            template = jinja2_env.get_template("fsh_vs_codes.j2")
            render_to_file(template, valueset, output_file)

import os
import re
from typing import List, Union

from openpyxl import load_workbook

from who_l3_smart_tools.utils import camel_case
from who_l3_smart_tools.utils.jinja2 import (
    DATA_TYPE_MAP,
    initalize_jinja_env,
    render_to_file,
)

jinja2_env = initalize_jinja_env(__file__)


# pylint: disable=too-few-public-methods
class QuestionnaireGenerator:
    """
    A class that generates FHIR Questionnaire resources from an Excel file.

    Args:
        input_file (str): The path to the input Excel file.
        output_dir (str): The directory where the generated FHIR Questionnaire
            resources will be saved.

    Attributes:
        input_file (str): The path to the input Excel file.
        output_dir (str): The directory where the generated FHIR Questionnaire
            resources will be saved.
        _activities (dict): A dictionary to store the activities and their associated
            questionnaire items.
        workbook: The loaded Excel workbook.

    Methods:
        generate_fsh_from_excel: Generates FHIR Questionnaire resources from the Excel file.
        _add_items_to_activity: Adds questionnaire items to the specified activity.

    """

    def __init__(self, input_file, output_dir):
        self.input_file = input_file
        self.output_dir = output_dir
        self._activities = {}
        self.workbook = load_workbook(self.input_file)

    def generate_fsh_from_excel(self):
        """
        Generates FHIR Questionnaire resources from the Excel file.

        This method iterates through each sheet in the workbook and extracts
            the questionnaire items.
        It then organizes the questionnaire items into activities and generates FHIR
            Questionnaire resources
        for each activity. The generated resources are saved in the specified
            output directory.

        """
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

        for sheet_name in self.workbook.sheetnames:
            if not re.match(r"HIV\.[A-Z\-]+\s", sheet_name):
                continue
            current_activity_id = None
            questionnaire_items = []

            sheet = self.workbook[sheet_name]
            header = None
            for row in sheet.iter_rows(values_only=True):
                if header is None:
                    header = row
                    continue
                row = dict(zip(header, row))
                activity_id = row["Activity ID"]

                # handle an activity change
                if isinstance(activity_id, str) and activity_id != current_activity_id:
                    # write out any existing activity
                    self._add_items_to_activity(
                        current_activity_id, questionnaire_items
                    )

                    # start a new activity
                    current_activity_id = activity_id
                    # NB The template gets formatted when written
                    questionnaire_items = []

                data_type = row["Data Type"]

                # we only want questions on the questionnaires
                if data_type == "Codes":
                    continue

                data_element_id = row["Data Element ID"]

                if not isinstance(data_element_id, str) or not data_element_id:
                    continue

                questionnaire_item = {
                    "data_element_id": data_element_id,
                    "data_element_label": str(row["Data Element Label"])
                    .replace("*", "")
                    .replace("[", "")
                    .replace("]", "")
                    .replace('"', "'")
                    .strip(),
                    "data_type": DATA_TYPE_MAP[data_type],
                    "required": "true" if str(row["Required"]) == "R" else "false",
                }

                # coded answers should be bound to a dataset
                if data_type == "Coding":
                    questionnaire_item["has_valueset"] = True

                questionnaire_items.append(questionnaire_item)

            self._add_items_to_activity(current_activity_id, questionnaire_items)

        for activity_code, activity in self._activities.items():
            _filename = os.path.join(self.output_dir, f"{activity_code}.fsh")
            render_to_file(
                jinja2_env.get_template("questionnaire.j2"),
                activity,
                _filename,
            )

    def _add_items_to_activity(
        self, current_activity_id: Union[str, None], questionnaire_items: List[str]
    ):
        """
        Adds questionnaire items to the specified activity.

        Args:
            current_activity_id (str or None): The ID of the current activity.
            questionnaire_items (list): The list of questionnaire items to be added.

        """
        if current_activity_id is not None:
            if "\n" in current_activity_id:
                activities = current_activity_id.split("\n")
            else:
                activities = [current_activity_id]

            for activity in activities:
                if " " in activity:
                    activity_code, activity_description = activity.split(" ", 1)
                    activity_desc_camel = camel_case(activity_description)
                    activity_desc_camel = (
                        activity_desc_camel[0].upper() + activity_desc_camel[1:]
                    )
                else:
                    activity_code = activity
                    activity_description = activity_desc_camel = activity.split(".", 1)[
                        1
                    ]

                if activity_code not in self._activities:
                    self._activities[activity_code] = {
                        "activity_id": f"{activity_code}{activity_desc_camel}",
                        "activity_title": activity_description,
                        "activity_title_description": activity_description[0].lower()
                        + activity_description[1:],
                        "questionnaire_items": [],
                    }

                self._activities[activity_code][
                    "questionnaire_items"
                ] += questionnaire_items

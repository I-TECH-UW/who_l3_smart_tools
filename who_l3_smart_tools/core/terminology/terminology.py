import csv
import json
from typing import Optional, Union

from openpyxl import load_workbook

from who_l3_smart_tools.core.terminology.schema import ConceptSchema


# pylint: disable=too-few-public-methods
class ConceptRow:
    """
    Represents a concept row.

    Args:
        row (dict): The row data.
        schema (type[ConceptSchema]): The schema for the concept.
        values_set (Union[dict, None]): The value set for the concept.
        **kwargs: Additional keyword arguments.

    Attributes:
        row (dict): The row data.
        schema (type[ConceptSchema]): The schema for the concept.
        converted_row (dict): The processed row data.
        value_set (Union[dict, None]): The value set for the concept.

    Methods:
        _format_name(name: str) -> dict:
            Formats the given name into a dictionary with specific keys.

        _format_description(description: str) -> dict:
            Formats the given description into a dictionary with specific keys.

        _process_row_required_fields(extra_args: dict) -> dict:
            Processes the required fields of the row.

        _process_value_set() -> None:
            Process the value set based on the datatype.

        _add_extras() -> None:
            Processes the row for JSON output.
    """

    def __init__(
        self,
        row: dict,
        schema: type[ConceptSchema],
        values_set: Union[dict, None],
        **kwargs,
    ) -> None:
        self.row = row
        self.schema = schema
        self.converted_row = None
        self.value_set = values_set
        self._process_row_required_fields(kwargs)
        self._add_extras()
        self._process_value_set()

    @staticmethod
    def _format_name(name: str) -> dict:
        """
        Formats the given name into a dictionary with specific keys.

        Args:
            name (str): The name to be formatted.

        Returns:
            dict: A dictionary containing the formatted name with the following keys:
                - "name": The original name.
                - "locale": The locale of the name (set to "en").
                - "locale_preferred": A boolean indicating if the name is preferred in the
                    locale (set to True).
                - "name_type": The type of the name (set to "Fully Specified").
        """
        return {
            "name": name,
            "locale": "en",
            "locale_preferred": True,
            "name_type": "Fully Specified",
        }

    @staticmethod
    def _format_description(description: str) -> dict:
        """
        Formats the given description into a dictionary with specific keys.

        Args:
            description (str): The description to be formatted.

        Returns:
            dict: A dictionary containing the formatted description with the following keys:
                - "description": The original description.
                - "locale": The locale of the description (set to "en").
                - "locale_preferred": A boolean indicating if the description is preferred in the
                    locale (set to True).
                - "description_type": The type of the description (set to "Definition").
        """
        return {
            "description": description,
            "locale": "en",
            "locale_preferred": True,
            "description_type": "Definition",
        }

    def _process_row_required_fields(self, extra_args: dict):
        """
        Processes the required fields of the row.

        Args:
            extra_args (dict): Additional arguments.

        Returns:
            dict: The processed row data.
        """
        converted_row = {
            "id": self.row.pop(self.schema.id),
            "name": self._format_name(self.row.pop(self.schema.name)),
            "datatype": self.row.pop(self.schema.datatype),
            "description": self._format_description(
                self.row.pop(self.schema.description)
            ),
        }
        for index, name in enumerate(self.schema.additional_names):
            converted_row[f"name[{index+1}]"] = self._format_name(self.row.pop(name))
        for index, description in enumerate(self.schema.additional_descriptions):
            converted_row[f"description[{index+1}]"] = self._format_description(
                self.row.pop(description)
            )
        converted_row.update(extra_args)
        self.converted_row = converted_row

    def _process_value_set(self):
        """
        Process the value set based on the datatype.

        If the datatype is 'codes', the method checks if a value set exists and raises a
        ValueError if not. It then appends a new coding to the value set.

        If the datatype is 'coding', the method creates a new value set with the specified
        system, coding, and text.

        If the datatype is neither 'codes' nor 'coding', the method sets the value set to None.

        """
        if self.converted_row.get("datatype").lower() == "codes":
            if not self.value_set:
                raise ValueError(
                    f"Value set is required for datatype 'codes'. {self.converted_row}"
                )
            codings = self.value_set["type"]["coding"]
            codings.append(
                {
                    "system": "http://hl7.org/fhir/v2/0203",
                    "code": self.converted_row["id"].replace(".", ""),
                    "display": self.converted_row["name"]["name"],
                }
            )
        elif self.converted_row.get("datatype").lower() == "coding":
            self.value_set = {
                "system": "https://fhir.staging.aws.openconceptlab.org",
                "type": {"coding": [], "text": self.converted_row["name"]},
                "value": "/orgs/DIGI/CodeSystem/HIVConcept",
            }
        else:
            self.value_set = None

    def _add_extras(self):
        """
        Processes the row for JSON output.
        """
        self.converted_row["extras"] = {}
        for key, value in self.row.items():
            if key in self.schema.exclude_columns:
                continue
            if self.schema.include_columns and key not in self.schema.include_columns:
                continue
            self.converted_row["extras"][key] = value


class ConceptResourceGenerator:
    """
    A class that generates concept resources from data files.

    Args:
        owner_id (str): The ID of the owner.
        file (str): The path to the data file.
        schema (type[ConceptSchema], optional): The schema to use for concept validation.
            Defaults to ConceptSchema.
        sheet_name_prefix (str, optional): The prefix of the sheet names to process.
            Defaults to "".

    Attributes:
        file (str): The path to the data file.
        schema (type[ConceptSchema]): The schema to use for concept validation.
        sheet_name_prefix (str): The prefix of the sheet names to process.
        extra_args (dict): Extra arguments for concept generation.
        valueset (Any): The current valueset being processed.
        valuesets (list): The list of valuesets collected.
        rows (list): The list of ConceptRow objects.

    Methods:
        collect_value_sets(row: dict) -> None:
            Collects value sets from the rows.
        _convert_rows() -> None:
            Converts the rows from the data files into ConceptRow objects.
        _format_csv_extras(key: dict) -> dict:
            Formats the extras for CSV output.
        _expand_objects_for_csv(key, obj: dict) -> dict:
            Expands the given dictionary object for CSV export.
        to_csv(path: str) -> None:
            Writes the rows to a CSV file.
        to_json(path: str) -> None:
            Writes the rows to a JSON file.
    """

    def __init__(
        self,
        owner_id: str,
        file: str,
        schema: type[ConceptSchema] = ConceptSchema,
        sheet_name_prefix: str = "",
    ):
        self.file = file
        self.schema = schema
        self.sheet_name_prefix = sheet_name_prefix
        self.extra_args = {
            "resource_type": "Concept",
            "owner_type": "Organization",
            "retired": False,
            "owner_id": owner_id,
            "concept_class": "Misc",
        }
        self.valueset = None
        self.valuesets = []

        self.rows = []
        self._convert_rows()

    def collect_value_sets(self, row: dict) -> None:
        """
        If we have a valueset, and the row datatype is not 'codes',
        we append the valueset to the valuesets list and reset the valueset.

        Args:
            row (dict): The row data.

        Returns:
            None
        """
        if self.valueset and row.converted_row["datatype"].lower() != "codes":
            self.valuesets.append(self.valueset)
            self.valueset = row.value_set
        if not self.valueset:
            self.valueset = row.value_set

    def _convert_rows(self) -> None:
        """
        Converts the rows from the data files into ConceptRow objects and
        appends them to the rows list. Also collects value sets from the rows.

        Returns:
            None
        """
        workbook = load_workbook(self.file)
        for sheet_name in workbook.sheetnames:
            if not sheet_name.startswith(self.sheet_name_prefix):
                continue
            sheet = workbook[sheet_name]
            header: Optional[list[str]] = None
            for row in sheet.iter_rows(values_only=True):
                # if header is set the current raw as the header and skip to the next row.
                if header is None:
                    header = row
                    continue
                row_as_dict = dict(zip(header, row))
                concept_row = ConceptRow(
                    row_as_dict,
                    self.schema,
                    self.valueset,
                    **self.extra_args,
                )
                self.rows.append(concept_row)
                self.collect_value_sets(concept_row)

    def _format_csv_extras(self, key: dict) -> dict:
        """
        Formats the extras for CSV output.

        Args:
            key (dict): The key to format.

        Returns:
            dict: The formatted key.
        """
        return {self.schema.format_extras_for_csv(k): v for k, v in key.items()}

    @staticmethod
    def _expand_objects_for_csv(key, obj: dict) -> dict:
        """
        Expands the given dictionary object for CSV export.

        Args:
            key (str): The key to be used for expanding the object.
            obj (dict): The dictionary object to be expanded.

        Returns:
            dict: The expanded dictionary object for CSV export.
        """
        expanded = {f"{key}_{k}": v for k, v in obj.items()}
        expanded[key] = expanded.pop(f"{key}_{key}")
        return expanded

    def to_csv(self, path: str) -> None:
        """
        Write the rows to a CSV file.

        Args:
            path (str): The path to the CSV file.

        Returns:
            None
        """
        with open(path, "w", encoding="utf-8") as output_file:
            if not self.rows:
                return
            writer = None
            for row in self.rows:
                csv_row = {}
                for key, value in row.converted_row.items():
                    if key == "extras":
                        csv_row.update(self._format_csv_extras(value))
                    elif isinstance(value, dict):
                        csv_row.update(self._expand_objects_for_csv(key, value))
                    elif key == "retired":
                        csv_row[key] = "true" if value else "false"
                    else:
                        csv_row[key] = value

                # if writer is None, we are on the first row and should write the header.
                if writer is None:
                    writer = csv.DictWriter(output_file, fieldnames=csv_row.keys())
                    writer.writeheader()
                writer.writerow(csv_row)

    def to_json(self, path: str) -> None:
        """
        Write the rows to a JSON file.

        Args:
            path (str): The path to the JSON file.

        Returns:
            None
        """
        with open(path, "w", encoding="utf-8") as output_file:
            json.dump(
                [row.converted_row for row in self.rows],
                output_file,
                indent=2,
            )
